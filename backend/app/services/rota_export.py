from datetime import datetime
from html import escape
from io import BytesIO
from typing import Any, List

from app.schemas import RotaDetailResponse, RotaSummaryRow

_STATUS_LABELS = {
    "on_time": "On time",
    "late": "Late",
    "absent": "Absent",
    "no_show": "No show",
}


def _fmt_day(dk: str) -> str:
    try:
        return datetime.strptime(dk, "%Y-%m-%d").strftime("%a %d %b")
    except ValueError:
        return dk


def _time_mins(t: str) -> int:
    parts = (t or "0:0").split(":")
    h = int(parts[0]) if parts and parts[0].isdigit() else 0
    m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
    return h * 60 + m


def _calc_hours(shift: dict[str, Any], incl_breaks: bool) -> float:
    sm = _time_mins(str(shift.get("start") or "0:00"))
    em = _time_mins(str(shift.get("end") or "0:00"))
    mins = em - sm
    if mins < 0:
        mins += 24 * 60
    if not incl_breaks:
        mins -= int(shift.get("breakH") or 0) * 60 + int(shift.get("breakM") or 0)
    return max(0.0, mins / 60.0)


def _shift_cell_html(shift: dict[str, Any], att: dict[str, Any] | None) -> str:
    site = (shift.get("site") or shift.get("notes") or "One-off").strip()
    lines = [f"{shift.get('start', '')}–{shift.get('end', '')}", escape(site[:36])]
    status = (att or {}).get("status")
    if status:
        lines.append(escape(_STATUS_LABELS.get(str(status), str(status).replace("_", " ").title())))
    note = ((att or {}).get("note") or "").strip()
    if note:
        lines.append(escape(note[:48]))
    return "<br/>".join(lines)


def export_planner_rota_pdf(data: dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    name = escape(str(data.get("rotaName") or "Rota"))
    days = list(data.get("days") or [])
    employees = list(data.get("employees") or [])
    shifts = data.get("shifts") or {}
    attendance = data.get("attendance") or {}
    incl_breaks = bool(data.get("inclBreaks", False))

    page = landscape(A4)
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=0.5 * cm,
        rightMargin=0.5 * cm,
        topMargin=0.7 * cm,
        bottomMargin=0.7 * cm,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=6.5,
        leading=8,
        wordWrap="CJK",
    )
    name_style = ParagraphStyle(
        "empName",
        parent=styles["Normal"],
        fontSize=7.5,
        leading=9,
        fontName="Helvetica-Bold",
        wordWrap="CJK",
    )
    header_style = ParagraphStyle("hdr", parent=styles["Normal"], fontSize=7, leading=8, fontName="Helvetica-Bold")

    story: list[Any] = [Paragraph(name, styles["Title"])]
    if days:
        story.append(
            Paragraph(
                f"{_fmt_day(days[0])} – {_fmt_day(days[-1])} · {len(days)} days · {len(employees)} employees",
                styles["Normal"],
            )
        )
    story.append(Spacer(1, 8))

    usable = page[0] - 1.0 * cm
    emp_col = 4.2 * cm  # wide enough for full names
    total_col = 1.2 * cm
    # Keep day columns readable; split across pages when too many days
    min_day = 1.6 * cm
    max_days_per_page = max(1, int((usable - emp_col - total_col) // min_day))

    def build_chunk(day_chunk: list[str], show_hours: bool) -> Table:
        hdr = [Paragraph("Employee", header_style)]
        for dk in day_chunk:
            hdr.append(Paragraph(escape(_fmt_day(dk)), header_style))
        if show_hours:
            hdr.append(Paragraph("Hours", header_style))
        table_data = [hdr]

        for emp in employees:
            emp_id = str(emp.get("id") or "")
            emp_name = escape(str(emp.get("name") or "").strip() or "—")
            row = [Paragraph(emp_name, name_style)]
            total_h = 0.0
            for dk in day_chunk:
                day_shifts = list((shifts.get(emp_id) or {}).get(dk) or [])
                if not day_shifts:
                    row.append(Paragraph("—", cell_style))
                    continue
                parts: list[str] = []
                for idx, sh in enumerate(day_shifts):
                    key = f"{emp_id}:{dk}:{idx}"
                    att = attendance.get(key)
                    parts.append(_shift_cell_html(sh, att))
                    total_h += _calc_hours(sh, incl_breaks)
                row.append(Paragraph("<br/><br/>".join(parts), cell_style))
            if show_hours:
                # Hours for whole rota (all days), not just this chunk
                full_h = 0.0
                for dk in days:
                    for sh in list((shifts.get(emp_id) or {}).get(dk) or []):
                        full_h += _calc_hours(sh, incl_breaks)
                row.append(Paragraph(f"{full_h:.1f}", cell_style))
            table_data.append(row)

        day_w = (usable - emp_col - (total_col if show_hours else 0)) / max(len(day_chunk), 1)
        col_widths = [emp_col] + [day_w] * len(day_chunk)
        if show_hours:
            col_widths.append(total_col)
        grid = Table(table_data, colWidths=col_widths, repeatRows=1)
        grid.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f8fafc")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return grid

    if not days:
        story.append(Paragraph("No days in this rota.", styles["Normal"]))
    else:
        chunks = [days[i : i + max_days_per_page] for i in range(0, len(days), max_days_per_page)]
        for i, chunk in enumerate(chunks):
            if i > 0:
                story.append(PageBreak())
                story.append(Paragraph(f"{name} (continued)", styles["Heading2"]))
                story.append(Spacer(1, 6))
            # Show hours column on the last chunk only
            story.append(build_chunk(chunk, show_hours=(i == len(chunks) - 1)))

    doc.build(story)
    return buf.getvalue()


def export_rota_xlsx(details: List[RotaDetailResponse], summary: List[RotaSummaryRow]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Shifts"
    hdr = [
        "ID",
        "Date",
        "Guard",
        "Site",
        "Client",
        "Start",
        "End",
        "Break (min)",
        "Type",
        "Hours",
        "Status",
        "Late (min)",
    ]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in details:
        ws.append(
            [
                d.id,
                str(d.date),
                d.guard_name,
                d.site_name,
                d.client_name or "",
                d.shift_start or "",
                d.shift_end or "",
                d.break_minutes,
                d.shift_type,
                d.hours,
                d.attendance_status,
                d.late_minutes or "",
            ]
        )
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Guard", "Total hours", "Late arrivals", "Committed (period)", "Overtime"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for s in summary:
        ws2.append([s.guard_name, s.total_hours, s.late_arrivals, s.committed_hours, s.overtime_hours])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rota_pdf(details: List[RotaDetailResponse], summary: List[RotaSummaryRow]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Rota — shifts", styles["Title"]))
    story.append(Spacer(1, 12))
    if details:
        data = [["Date", "Guard", "Site", "Client", "Shift", "Type", "Hrs", "Status"]]
        for d in details:
            sh = f"{d.shift_start or '-'}–{d.shift_end or '-'}"
            data.append(
                [
                    str(d.date),
                    d.guard_name,
                    d.site_name[:24],
                    (d.client_name or "")[:20],
                    sh,
                    d.shift_type,
                    f"{d.hours:.1f}",
                    d.attendance_status,
                ]
            )
        t = Table(data, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(t)
    story.append(Spacer(1, 24))
    story.append(Paragraph("Summary by guard", styles["Title"]))
    story.append(Spacer(1, 12))
    if summary:
        sdata = [["Guard", "Total hrs", "Late", "Committed", "Overtime"]]
        for s in summary:
            sdata.append(
                [
                    s.guard_name[:30],
                    f"{s.total_hours:.1f}",
                    str(s.late_arrivals),
                    f"{s.committed_hours:.1f}",
                    f"{s.overtime_hours:.1f}",
                ]
            )
        t2 = Table(sdata, repeatRows=1)
        t2.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        story.append(t2)
    doc.build(story)
    return buf.getvalue()
